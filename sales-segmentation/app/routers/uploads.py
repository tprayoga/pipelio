"""
Upload & impor data pipeline dari Excel/CSV.

Menggantikan integrasi Base44 (UploadFile + ExtractDataFromUploadedFile).
Berbeda dengan versi Base44 yang memakai LLM untuk menebak isi file, di sini
pemetaan kolomnya eksplisit dan bisa diperiksa — untuk data operasional kantor,
hasil yang dapat diprediksi lebih penting daripada fleksibilitas menebak.
"""

from __future__ import annotations

import math
from io import BytesIO
from typing import Any, Dict, List

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app import schemas
from app.db import get_db
from app.deps import require_master_data
from app.models import Pipeline, User

router = APIRouter(prefix="/uploads", tags=["uploads"])

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MAX_UPLOAD_BYTES = 10 * 1024 * 1024   # 10 MB
MAX_PREVIEW_ROWS = 50
MAX_IMPORT_ROWS = 20_000

# Nama kolom bebas -> field Pipeline. Dibandingkan setelah lowercase & trim.
COLUMN_ALIASES = {
    "customer": "customer", "pelanggan": "customer", "nama customer": "customer",
    "company": "company", "perusahaan": "company",
    "industry": "industry", "industri": "industry",
    "project_name": "project_name", "project name": "project_name",
    "nama proyek": "project_name", "proyek": "project_name",
    "project_category": "project_category", "project category": "project_category",
    "kategori": "project_category",
    "estimated_value": "estimated_value", "estimated value": "estimated_value",
    "nilai": "estimated_value", "nilai proyek": "estimated_value",
    "estimated_closing": "estimated_closing", "estimated closing": "estimated_closing",
    "tanggal closing": "estimated_closing",
    "pic": "pic", "email": "email", "phone": "phone", "telepon": "phone",
    "notes": "notes", "catatan": "notes",
    "stage": "stage", "tahap": "stage",
    "sales_name": "sales_name", "sales name": "sales_name",
    "sales executive": "sales_name", "sales": "sales_name", "nama sales": "sales_name",
    "quarter": "quarter", "kuartal": "quarter",
    "year": "year", "tahun": "year",
    "status": "status",
}

STAGE_ALIASES = {
    "leads": "Leads", "lead": "Leads",
    "prospecting": "Prospecting", "prospect": "Prospecting",
    "negotiating": "Negotiating", "negotiation": "Negotiating", "negosiasi": "Negotiating",
    "won": "Won", "win": "Won", "menang": "Won",
    "lost": "Lost", "lose": "Lost", "kalah": "Lost",
    "hold": "Hold", "ditunda": "Hold",
}

REQUIRED_FIELDS = ("customer", "project_name")


def _clean(value: Any) -> Any:
    """NaN/NaT pandas tidak bisa dikirim sebagai JSON — ubah jadi None."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if value is pd.NaT:
        return None
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    return value


def _parse_file(upload: UploadFile, content: bytes) -> pd.DataFrame:
    name = (upload.filename or "").lower()
    try:
        if name.endswith(".csv"):
            return pd.read_csv(BytesIO(content))
        if name.endswith((".xlsx", ".xls")):
            return pd.read_excel(BytesIO(content))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"File tidak bisa dibaca: {exc}")
    raise HTTPException(status_code=400, detail="Format harus .xlsx, .xls, atau .csv.")


def _normalize(df: pd.DataFrame) -> tuple:
    """Petakan kolom ke skema Pipeline. Kembalikan (rows, warnings)."""
    warnings: List[str] = []

    rename, unknown = {}, []
    for column in df.columns:
        key = str(column).strip().lower()
        if key in COLUMN_ALIASES:
            rename[column] = COLUMN_ALIASES[key]
        else:
            unknown.append(str(column))
    df = df.rename(columns=rename)

    if unknown:
        warnings.append("Kolom diabaikan karena tidak dikenali: " + ", ".join(unknown))

    missing = [f for f in REQUIRED_FIELDS if f not in df.columns]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Kolom wajib tidak ditemukan: " + ", ".join(missing)
                + ". Minimal dibutuhkan kolom Customer dan Project Name."
            ),
        )

    known = [c for c in df.columns if c in set(COLUMN_ALIASES.values())]
    df = df[known]

    if "stage" in df.columns:
        unmapped = set()

        def map_stage(value):
            if value is None or (isinstance(value, float) and math.isnan(value)):
                return "Leads"
            mapped = STAGE_ALIASES.get(str(value).strip().lower())
            if mapped is None:
                unmapped.add(str(value))
                return "Leads"
            return mapped

        df["stage"] = df["stage"].map(map_stage)
        if unmapped:
            warnings.append(
                "Nilai stage tidak dikenali, dijadikan 'Leads': " + ", ".join(sorted(unmapped))
            )

    for numeric in ("estimated_value", "year"):
        if numeric in df.columns:
            df[numeric] = pd.to_numeric(df[numeric], errors="coerce")

    if "quarter" in df.columns:
        df["quarter"] = (
            df["quarter"].astype(str).str.strip().str.upper()
            .where(lambda s: s.isin(["Q1", "Q2", "Q3", "Q4"]), None)
        )

    rows: List[Dict[str, Any]] = []
    dropped = 0
    for record in df.to_dict(orient="records"):
        cleaned = {k: _clean(v) for k, v in record.items()}
        # Baris tanpa customer/project tidak bisa disimpan — biasanya baris
        # kosong di akhir sheet.
        if not cleaned.get("customer") or not cleaned.get("project_name"):
            dropped += 1
            continue
        cleaned.setdefault("stage", "Leads")
        if cleaned.get("estimated_value") is None:
            cleaned["estimated_value"] = 0
        if cleaned.get("year") is not None:
            cleaned["year"] = int(cleaned["year"])
        rows.append(cleaned)

    if dropped:
        warnings.append(f"{dropped} baris dilewati karena Customer atau Project Name kosong.")

    return rows, warnings


# Kolom pada template. Judulnya memakai penamaan yang lazim dipakai staf, bukan
# nama field internal — semuanya sudah terdaftar di COLUMN_ALIASES.
TEMPLATE_COLUMNS = [
    ("Customer", True, "Nama perusahaan pelanggan"),
    ("Project Name", True, "Nama proyek atau pengadaan"),
    ("Industry", False, "Sektor pelanggan, mis. Manufaktur"),
    ("Stage", False, f"Salah satu: {', '.join(dict.fromkeys(STAGE_ALIASES.values()))}"),
    ("Sales Executive", False, "Nama sales penanggung jawab"),
    ("Quarter", False, "Q1, Q2, Q3, atau Q4"),
    ("Year", False, "Tahun, mis. 2025"),
    ("Nilai", False, "Nilai proyek dalam rupiah, angka saja tanpa titik"),
    ("Tanggal Closing", False, "Perkiraan tanggal closing, format YYYY-MM-DD"),
    ("PIC", False, "Nama kontak di pihak pelanggan"),
    ("Email", False, "Email kontak"),
    ("Phone", False, "Telepon kontak"),
    ("Catatan", False, "Keterangan tambahan"),
]

TEMPLATE_EXAMPLES = [
    ["PT Sinar Baja Elektrik", "Pengadaan Server Rack", "Manufaktur", "Won",
     "Ari Ardiansyah Habib", "Q1", 2025, 250000000, "2025-03-14",
     "Andi Wijaya", "andi@sinarbaja.co.id", "021-5551234", ""],
    ["PT Telkom Akses", "Implementasi Jaringan Fiber", "Telekomunikasi", "Negotiating",
     "Usep Kurnia", "Q2", 2025, 480000000, "2025-05-20",
     "Rina Kusuma", "rina@telkomakses.co.id", "021-5555678", "Menunggu revisi RAB"],
]


def _build_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Data Pipeline"

    header_fill = PatternFill("solid", fgColor="122E61")
    required_fill = PatternFill("solid", fgColor="C0392B")

    for index, (name, required, _) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        # Kolom wajib diberi warna berbeda supaya terlihat sebelum orang
        # membuka sheet petunjuk.
        cell.fill = required_fill if required else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = max(len(name) + 4, 16)

    for row_index, example in enumerate(TEMPLATE_EXAMPLES, start=2):
        for col_index, value in enumerate(example, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    sheet.freeze_panes = "A2"

    guide = workbook.create_sheet("Petunjuk")
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 12
    guide.column_dimensions["C"].width = 70

    intro = guide.cell(row=1, column=1, value="Cara mengisi template")
    intro.font = Font(bold=True, size=13)

    notes = [
        "Isi mulai baris ke-2 pada sheet 'Data Pipeline'. Dua baris contoh boleh dihapus.",
        "Satu baris = satu deal. Kolom bertanda WAJIB tidak boleh kosong; barisnya akan dilewati.",
        "Urutan kolom bebas. Kolom yang tidak dikenali diabaikan dan dilaporkan saat pratinjau.",
        "Nilai Stage di luar daftar akan dijadikan 'Leads' disertai peringatan.",
        "Periksa hasil pratinjau di aplikasi sebelum menekan tombol Import.",
    ]
    for offset, note in enumerate(notes):
        guide.cell(row=3 + offset, column=1, value=f"• {note}")

    table_start = 3 + len(notes) + 1
    for column, title in enumerate(["Kolom", "Wajib?", "Keterangan"], start=1):
        cell = guide.cell(row=table_start, column=column, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for offset, (name, required, description) in enumerate(TEMPLATE_COLUMNS, start=1):
        row = table_start + offset
        guide.cell(row=row, column=1, value=name)
        wajib = guide.cell(row=row, column=2, value="WAJIB" if required else "opsional")
        if required:
            wajib.font = Font(bold=True, color="C0392B")
        guide.cell(row=row, column=3, value=description)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@router.get("/template")
def download_template(_: User = Depends(require_master_data)):
    """Template Excel berisi kolom yang dikenali beserta contoh pengisian."""
    return StreamingResponse(
        BytesIO(_build_template()),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": 'attachment; filename="template_pipeline_pipelio.xlsx"'
        },
    )


@router.post("/preview", response_model=schemas.UploadPreview)
async def preview(
    file: UploadFile = File(...),
    _: User = Depends(require_master_data),
):
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Ukuran file melebihi 10 MB.")
    if not content:
        raise HTTPException(status_code=400, detail="File kosong.")

    df = _parse_file(file, content)
    if df.empty:
        raise HTTPException(status_code=400, detail="File tidak berisi baris data.")
    if len(df) > MAX_IMPORT_ROWS:
        raise HTTPException(
            status_code=413, detail=f"Maksimal {MAX_IMPORT_ROWS:,} baris per file."
        )

    rows, warnings = _normalize(df)
    if not rows:
        raise HTTPException(
            status_code=400, detail="Tidak ada baris valid setelah pemeriksaan."
        )

    return schemas.UploadPreview(
        rows=rows[:MAX_PREVIEW_ROWS],
        total=len(rows),
        columns=list(rows[0].keys()),
        warnings=warnings,
    )


@router.post("/import", response_model=schemas.ImportResult)
def import_rows(
    payload: schemas.ImportRequest,
    db: Session = Depends(get_db),
    _: User = Depends(require_master_data),
):
    if not payload.rows:
        raise HTTPException(status_code=400, detail="Tidak ada baris untuk diimpor.")
    if len(payload.rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=413, detail=f"Maksimal {MAX_IMPORT_ROWS:,} baris.")

    # Satu transaksi: bila ada satu baris gagal, tidak ada yang masuk separuh.
    try:
        db.add_all([Pipeline(**row.model_dump()) for row in payload.rows])
        db.commit()
    except Exception:
        db.rollback()
        raise

    return schemas.ImportResult(imported=len(payload.rows))
